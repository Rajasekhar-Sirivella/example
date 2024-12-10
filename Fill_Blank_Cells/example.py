import streamlit as st
import openpyxl
from crewai import Agent, Task, Crew
# Importing crewAI tools
from crewai_tools import SerperDevTool, JSONSearchTool
import os

os.environ["SERPER_API_KEY"] = "Your Key"
search_tool = SerperDevTool()
json_tool = JSONSearchTool()
def convert_to_string(data):
    result = ""
    for row in data:
        row_str = "\t".join([str(item) if item is not None else "null" for item in row])
        result += row_str + "\n"
    return result
#Extracting data from excel sheet
def extract_from_excel(excel_file):
    workbook=openpyxl.load_workbook(excel_file)
    data_list=[]
    for sheet in workbook.sheetnames:
        worksheet=workbook[sheet]
        data=[]
        for row in worksheet.iter_rows(values_only=True):
            data.append(row)
        data_list.append(convert_to_string(data))
        
    # st.write(data_list)
    # return "\n".join(str(item) for item in data_list)
    return data_list

st.title("Fill Blank Cells")

# File upload
uploaded_file = st.file_uploader("Upload your Runbook (EXCEL)", type="xlsx")

if uploaded_file is not None:
    with st.spinner("Extracting text from the excel..."):
        data_list = extract_from_excel(uploaded_file)

        # Display extracted text if available
        if data_list:
            st.write("### Extracted Text:")
            st.write(data_list)

# Create agents
researcher = Agent(
    role='Research Analyst',
    goal='Provide correct suitable value for null value',
    backstory='An expert analyst with seaching the values',
    tools=[search_tool],
    verbose=True
)

writer = Agent(
    role='json Writer',
    goal='provide json data with given data',
    backstory='A skilled json writer with data',
    tools=[json_tool],
    verbose=True
)

# Define tasks
research = Task(
    description='Research the correct suitable value for null value',
    expected_output='filled null values with correct values, not get the values leave as it is as null.',
    agent=researcher
)

write = Task(
    description='write the response in the format of json',
    expected_output='json resonse with all correct values',
    agent=writer,
    output_file="report.md"
)

# Assemble a crew with planning enabled
crew = Crew(
    agents=[researcher, writer],
    tasks=[research, write],
    verbose=True,
    planning=True,  # Enable planning feature
)

# Execute tasks
resp=crew.kickoff(inputs=data_list[0])
st.write(resp)
