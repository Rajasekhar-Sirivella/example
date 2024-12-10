import streamlit as st
import openpyxl
from langchain.agents import AgentType, initialize_agent
from langchain_community.utilities import GoogleSerperAPIWrapper
from langchain_core.tools import Tool
from langchain.agents import AgentExecutor, create_tool_calling_agent
from langchain_google_vertexai.chat_models import ChatVertexAI
from google.cloud import aiplatform
from langchain.prompts import PromptTemplate
from langchain_community.tools.google_serper import GoogleSerperResults
from langgraph.prebuilt import create_react_agent
from langchain_core.messages import SystemMessage, HumanMessage
import os
 
PROJECT_ID = "tcs-cto-retail"
aiplatform.init(project=PROJECT_ID, location="us-central1")

os.environ["SERPER_API_KEY"] = "3d5b748c4e819a8811d2a77c651fbc8a940a8b1e"
llm = ChatVertexAI(model="gemini-1.5-flash-002")
search = GoogleSerperAPIWrapper()

#you are a data analyst, analyze the provided data and got any null values find out the correct value for null value or else put same as null. 
        # return the response in the format of json.
        # input : {input}
        # {agent_scratchpad}
# prompt=PromptTemplate(input_variables=["input", "agent_scratchpad"],
    # template="""
    #     You are a highly skilled and detail-oriented excel data analyst. Your primary responsibility is to analyze the provided excel data(string format) row by row thoroughly and address any null or missing values. Here are the steps you must follow:
    #     1. **Null Value Identification**:
    #     - Carefully examine the provided excel data(string format) to identify any null, missing, or incomplete values.
    #     2. **Handling Null Values**:
    #     - For each null value:
    #         - Search that row with help of provided tool and get correct value
    #         - If you can determine a correct value based on context or logical inference, replace the null value with the determined value.
    #         - If it is not possible to determine the correct value, retain the null value and indicate it as `"null"` in the response.
    #     3. **Data Integrity**:
    #     - Ensure that all existing values in the dataset remain unchanged. Do not alter any values other than nulls.
    #     4. **Response Format**:
    #     - Return the analyzed and updated dataset in a structured JSON format. Ensure the JSON is well-formatted and valid.
    #     input : {input}
    #     {agent_scratchpad}
    # """)
prompt=PromptTemplate(input_variables=["input", "agent_scratchpad"],
    template="""
    Role and Task:
    You are an expert Excel data analyst specialized in handling datasets with missing or incomplete values. Your task is to thoroughly examine the provided dataset (in string format) row by row, 
    identify null or missing values, and address them systematically using the column headers for context and logical inference.
    
    Steps to Follow:

        1.Null Value Identification:
        * Parse the provided Excel data (in string format) into a structured tabular representation using the provided headers.
        * Examine each row and column to identify any null, missing, or incomplete values (e.g., blank cells, null, or placeholders like NA).

        2.Handling Null Values:
        * For every null or missing value encountered:
            * Use the provided headers to understand the type of data expected in that column.
            * Use the provided tool to fetch or infer the correct value for that specific cell, considering the context of the column (e.g., numerical, categorical, date).
            * If a correct value can be logically inferred or determined based on the column header and other data in the same row, replace the null value with it.
            * If it is not possible to fetch or infer the value, explicitly retain it as "null".

        3. Data Integrity:
        * Preserve the integrity of all non-null values in the dataset; do not modify them.
        * Ensure the final dataset remains consistent with the original structure and formatting.
        
        4.Response Format:
        * Return the processed dataset in a structured JSON format. 
        * The JSON must use the provided headers as keys and include updated rows with all null values handled.
        
    Input Format:
    * The input includes a string representing the Excel data and a set of headers.
    * The headers are explicitly defined to help identify and contextualize each column in the dataset.
    
    Output Requirements:
    * The response must include a structured and valid JSON format, adhering to the schema mentioned above.
    * Ensure that null values are updated accurately based on the column context provided by the headers.

    headers: {headers}
    input : {input}
    {agent_scratchpad}
    """)

# tools = [
#     Tool(
#         name="Google Search Tool",
#         func=search.run,
#         description="find out the correct null value, if not retun null value as same",
#     )
# ]


tools=[GoogleSerperResults()]

agent = create_tool_calling_agent(llm, tools, prompt)
agent_executor = AgentExecutor(agent=agent, tools=tools, verbose=True)


def convert_to_string(data):
    result = ""
    for row in data:
        row_str = ", ".join([str(item) if item is not None else "null" for item in row])
        result = result+"[" +row_str+"]" + "\n"
    return result
#Extracting data from excel sheet
def extract_from_excel(excel_file):
    workbook=openpyxl.load_workbook(excel_file)
    data_list=[]
    headers=[]
    for sheet in workbook.sheetnames:
        worksheet=workbook[sheet]
        data=[]
        for row in worksheet.iter_rows(values_only=True):
            data.append(row)
        headers.append("["+"".join(str(item) for item in data[0])+"]")
        data_list.append(convert_to_string(data[1:]))
        
    # st.write(data_list)
    # return "\n".join(str(item) for item in data_list)
    return headers, data_list

st.title("Fill Blank Cells")

# File upload
uploaded_file = st.file_uploader("Upload your (EXCEL)", type="xlsx")

if uploaded_file is not None:
    with st.spinner("Extracting text from the excel..."):
        headers, data_list = extract_from_excel(uploaded_file)

        # Display extracted text if available
        # if data_list:
        #     st.write("### Extracted Text:")
        st.write(data_list[0])
        #input="Product Name: LG OLED55C1PUB, Category: TV, Brand: LG, Model: OLED55C1PUB, Price($): 1799, Screen Size (inches): 4K, Resolution: UHD, Storage Capacity (GB): NULL, RAM (GB): NULL, Quality Rating (1-5): 5, Release Year: 2,021"
        input="Product Name: Samsung Galaxy S21 Ultra, Category: Mobile, Brand: Samsung, Model: SM-G998U, Price($): NULL, Screen Size (inches): 6.80, Resolution: 3200 x 1440, Storage Capacity (GB): 128, RAM (GB): 12, Quality Rating (1-5): 4.9, Release Year: 2,021"
        response=agent_executor.invoke({"headers":headers[0] ,"input": data_list[0]})
        st.write(response["output"])
        
