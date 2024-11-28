import streamlit as st
from langchain_groq import ChatGroq
import pandas as pd
from langchain.prompts import PromptTemplate
from langchain.text_splitter import RecursiveCharacterTextSplitter
from dotenv import load_dotenv
import os
from langchain.chains.summarize import load_summarize_chain
import openpyxl

load_dotenv()

os.environ['GROQ_API_KEY'] = os.getenv('GROQ_API_KEY')

#
def convert_to_string(data):
    result = ""
    for row in data:
        row_str = "\t".join([str(item) if item is not None else "" for item in row])
        result += row_str + "\n"
    return result
#Extracting data from excel sheet
def extract_from_excel(excel_file):
    workbook=openpyxl.load_workbook(excel_file)
    data_list=[]
    for sheet in workbook.sheetnames:
        worksheet=workbook[sheet]
        data=[]
        for row in worksheet.iter_rows(values_only=True):
            data.append(row)
        data_list.append(convert_to_string(data))
    st.write(data_list)
    return "\n".join(str(item) for item in data_list)

def summarize_runbook(runbook_text):
    # Split the text into smaller chunks (if it's long)
    text_splitter = RecursiveCharacterTextSplitter(
        chunk_size=1000, 
        chunk_overlap=50
    )
    # Split runbook text into Document objects
    split_docs = text_splitter.create_documents([runbook_text])
    
    # Initial concise summary prompt template
    # prompt_template = """Write headings at start if present after that, 
    # Write a concise summary of the following Test Cases:
    # {text}
    # CONCISE SUMMARY:"""
    prompt_template = """
    For each page, if headings are present, write them at the beginning. 
    Then, write a detailed but concise summary of the following Test Cases:
    {text}
    DETAILED AND CONCISE SUMMARY:
    """
    prompt = PromptTemplate.from_template(prompt_template)
    
    # Initialize LLM (LangChain with ChatGroq)
    llm = ChatGroq(temperature=0, model_name="llama-3.1-8b-instant")

    # Load the refine chain
    chain = load_summarize_chain(
        llm=llm,
        chain_type="refine",
        question_prompt=prompt,
        return_intermediate_steps=True,
        input_key="input_documents",
        output_key="output_text",
        verbose=True
    )

    # Run the chain
    result = chain({"input_documents": split_docs}, return_only_outputs=True)

    return result['output_text']

# Streamlit session state
if 'runbook_text' not in st.session_state:
    st.session_state.runbook_text = None
if 'summary' not in st.session_state:
    st.session_state.summary = None

st.title("Summarizer")

# File upload
uploaded_file = st.file_uploader("Upload your Runbook (EXCEL)", type="xlsx")

if uploaded_file is not None:
    with st.spinner("Extracting text from the excel..."):
        runbook_text = extract_from_excel(uploaded_file)
        st.session_state.runbook_text = runbook_text  # Store extracted text in session state

# Display extracted text if available
if st.session_state.runbook_text:
    st.write("### Extracted Text:")
    st.write(st.session_state.runbook_text[:500])  # Display first 500 characters as a preview

# Summarize button
if st.session_state.runbook_text and st.button("Summarize Runbook"):
    with st.spinner("Summarizing the runbook..."):
        summary = summarize_runbook(st.session_state.runbook_text)
        st.session_state.summary = summary  # Store summary in session state

# Display summary if available
if st.session_state.summary:
    st.write("### Summary:")
    st.write(st.session_state.summary)
